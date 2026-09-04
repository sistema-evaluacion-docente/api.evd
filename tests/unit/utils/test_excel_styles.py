"""Tests for the shared Excel styling helpers."""

from openpyxl import Workbook

from api.utils.excel_styles import (
    C_GREEN,
    C_ORANGE,
    C_RED,
    C_WHITE,
    bold_font,
    center,
    fill,
    left,
    score_color,
    style_data_row,
    style_section_header,
    style_table_header,
    style_title_row,
    thin_border,
)


class TestScoreColor:
    """score_color buckets a numeric score into a traffic-light color."""

    def test_none_is_gray(self):
        assert score_color(None) == "888888"

    def test_excellent_score_is_green(self):
        assert score_color(4.5) == C_GREEN

    def test_good_score_is_a_lighter_green(self):
        assert score_color(4.0) == "2E7D32"

    def test_borderline_score_is_orange(self):
        assert score_color(3.5) == C_ORANGE

    def test_low_score_is_red(self):
        assert score_color(2.0) == C_RED


class TestBuildingBlocks:
    """fill/bold_font/thin_border/center/left return configured openpyxl styles."""

    def test_fill_uses_the_given_color(self):
        result = fill("FF0000")

        assert result.fgColor.rgb == "00FF0000"

    def test_bold_font_defaults_to_the_given_color(self):
        result = bold_font(size=12, color="123456")

        assert result.bold is True
        assert result.size == 12
        assert result.color.rgb == "00123456"

    def test_bold_font_white_overrides_the_color(self):
        result = bold_font(white=True)

        assert result.color.rgb == f"00{C_WHITE}"

    def test_thin_border_sets_all_four_sides(self):
        result = thin_border()

        assert result.left.style == "thin"
        assert result.right.style == "thin"
        assert result.top.style == "thin"
        assert result.bottom.style == "thin"

    def test_center_alignment(self):
        result = center()

        assert result.horizontal == "center"
        assert result.wrap_text is True

    def test_left_alignment(self):
        result = left()

        assert result.horizontal == "left"
        assert result.wrap_text is True


class TestRowStyling:
    """The style_* helpers write and format cells on a real worksheet."""

    def test_style_title_row_merges_and_writes_the_text(self):
        ws = Workbook().active

        style_title_row(ws, 1, "Reporte", ncols=3)

        assert ws.cell(row=1, column=1).value == "Reporte"
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.merged_cells.ranges

    def test_style_section_header_merges_and_writes_the_text(self):
        ws = Workbook().active

        style_section_header(ws, 2, "Sección", ncols=3)

        assert ws.cell(row=2, column=1).value == "Sección"
        assert ws.merged_cells.ranges

    def test_style_table_header_writes_every_column(self):
        ws = Workbook().active

        style_table_header(ws, 3, ["Nombre", "Nota"])

        assert ws.cell(row=3, column=1).value == "Nombre"
        assert ws.cell(row=3, column=2).value == "Nota"
        assert ws.cell(row=3, column=1).border.left.style == "thin"

    def test_style_data_row_alternates_background(self):
        ws = Workbook().active

        style_data_row(ws, 4, ["Ana", 4.5], alternate=True)

        assert ws.cell(row=4, column=1).value == "Ana"
        assert ws.cell(row=4, column=1).fill.fgColor.rgb == "00F4F8FF"

    def test_style_data_row_default_background(self):
        ws = Workbook().active

        style_data_row(ws, 4, ["Ana", 4.5])

        assert ws.cell(row=4, column=1).fill.fgColor.rgb == f"00{C_WHITE}"

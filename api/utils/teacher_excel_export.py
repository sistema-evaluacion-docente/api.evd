"""
Excel export for a teacher's evaluation detail report.
"""

import io

import openpyxl
from fastapi.responses import StreamingResponse

from api.utils.excel_styles import (
    C_DIM_HDR,
    C_GREEN,
    C_RED,
    EXCEL_MIME,
    bold_font,
    fill,
    left,
    score_color,
    style_data_row,
    style_section_header,
    style_table_header,
    style_title_row,
    thin_border,
    center,
)
from openpyxl.styles import Alignment, Font


def build_teacher_report(
    detail: dict,
    comparison: dict | None,
    comments_by_course: dict[str, list],
) -> tuple[io.BytesIO, str]:
    """Build the teacher evaluation Excel workbook.

    Returns (buffer, filename).
    """

    dept_by_dim: dict[str, float | None] = {}
    dept_by_code: dict[str, float | None] = {}
    dept_name: str | None = None
    dept_overall: float | None = None

    if comparison:
        dept_name = comparison.get("department_name")
        dept_overall = comparison.get("department_overall_average")
        for dim in comparison["dimensions"]:
            dept_by_dim[dim["dimension"]] = dim["department_average"]
            for q in dim["questions"]:
                dept_by_code[q["code"]] = q["department_average"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Docente"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    r = 1

    style_title_row(ws, r, "REPORTE DE EVALUACIÓN DOCENTE — UFPS")
    ws.row_dimensions[r].height = 28
    r += 1
    r += 1

    style_section_header(ws, r, "INFORMACIÓN DEL DOCENTE")
    r += 1

    info_rows = [
        ("Nombre", detail["name"] or "—"),
        ("Código", detail["institutional_code"]),
        ("Tipo contrato", detail["contract_type"] or "—"),
        ("Período", detail["period_code"] or "—"),
    ]
    for label, value in info_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = bold_font()
        lc.fill = fill(C_DIM_HDR)
        lc.border = thin_border()
        lc.alignment = left()
        vc = ws.cell(row=r, column=2, value=value)
        vc.border = thin_border()
        vc.alignment = left()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    avg = detail["overall_average"]
    lc = ws.cell(row=r, column=1, value="Promedio docente")
    lc.font = bold_font()
    lc.fill = fill(C_DIM_HDR)
    lc.border = thin_border()
    lc.alignment = left()
    vc = ws.cell(row=r, column=2, value=f"{avg} / 5.0" if avg else "—")
    vc.font = Font(bold=True, color=score_color(avg))
    vc.border = thin_border()
    vc.alignment = left()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    lc = ws.cell(row=r, column=1, value="Promedio departamento")
    lc.font = bold_font()
    lc.fill = fill(C_DIM_HDR)
    lc.border = thin_border()
    lc.alignment = left()
    dept_avg_label = f"{dept_overall} / 5.0" if dept_overall else "N/D"
    if dept_name:
        dept_avg_label = (
            f"{dept_overall} / 5.0  ({dept_name})"
            if dept_overall
            else f"N/D  ({dept_name})"
        )
    vc = ws.cell(row=r, column=2, value=dept_avg_label)
    vc.font = Font(bold=True, color=score_color(dept_overall))
    vc.border = thin_border()
    vc.alignment = left()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    r += 1

    style_section_header(ws, r, "RESUMEN POR DIMENSIÓN")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Docente", "Departamento", "Diferencia"])
    r += 1
    for i, dim in enumerate(detail["dimensions"]):
        dim_avg = dim["average"]
        dept_dim_avg = dept_by_dim.get(dim["dimension"])
        diff = (
            round(dim_avg - dept_dim_avg, 2)
            if dim_avg is not None and dept_dim_avg is not None
            else None
        )
        diff_str = (
            f"+{diff}"
            if diff is not None and diff > 0
            else str(diff) if diff is not None else "N/D"
        )
        style_data_row(
            ws,
            r,
            [
                dim["dimension"],
                f"{dim_avg} / 5.0" if dim_avg is not None else "—",
                f"{dept_dim_avg} / 5.0" if dept_dim_avg is not None else "N/D",
                diff_str,
            ],
            alternate=i % 2 == 1,
        )
        ws.cell(row=r, column=2).font = Font(bold=True, color=score_color(dim_avg))
        ws.cell(row=r, column=3).font = Font(bold=True, color=score_color(dept_dim_avg))
        ws.cell(row=r, column=4).font = Font(
            bold=True, color=C_GREEN if (diff or 0) >= 0 else C_RED
        )
        r += 1
    r += 1

    sorted_dims = sorted(
        [d for d in detail["dimensions"] if d["average"] is not None],
        key=lambda d: d["average"],
        reverse=True,
    )
    style_section_header(ws, r, "FORTALEZAS — Top 2 dimensiones")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Promedio", "", ""])
    r += 1
    for dim in sorted_dims[:2]:
        style_data_row(ws, r, [dim["dimension"], f"{dim['average']} / 5.0", "", ""])
        ws.cell(row=r, column=2).font = Font(bold=True, color=C_GREEN)
        r += 1
    r += 1

    style_section_header(ws, r, "OPORTUNIDADES DE MEJORA — Top 2 dimensiones")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Promedio", "", ""])
    r += 1
    for dim in sorted_dims[-2:]:
        style_data_row(ws, r, [dim["dimension"], f"{dim['average']} / 5.0", "", ""])
        ws.cell(row=r, column=2).font = Font(bold=True, color=C_RED)
        r += 1
    r += 1

    style_title_row(ws, r, "MATRIZ DE EVALUACIÓN — 22 ÍTEMS")
    r += 1

    for dim in detail["dimensions"]:
        r += 1
        style_section_header(
            ws,
            r,
            f"{dim['dimension'].upper()}   |   Promedio: {dim['average']} / 5.0",
        )
        r += 1
        style_table_header(ws, r, ["Código", "Descripción", "Docente", "Departamento"])
        r += 1
        for i, q in enumerate(dim.get("questions", [])):
            dept_q_avg = dept_by_code.get(q["code"])
            style_data_row(
                ws,
                r,
                [
                    q["code"],
                    q["text"],
                    f"{q['score']} / 5.0" if q["score"] is not None else "—",
                    f"{dept_q_avg} / 5.0" if dept_q_avg is not None else "N/D",
                ],
                alternate=i % 2 == 1,
            )
            ws.cell(row=r, column=3).font = Font(
                bold=True, color=score_color(q["score"])
            )
            ws.cell(row=r, column=3).alignment = center()
            ws.cell(row=r, column=4).font = Font(
                bold=True, color=score_color(dept_q_avg)
            )
            ws.cell(row=r, column=4).alignment = center()
            r += 1

    if comments_by_course:
        r += 1
        style_title_row(ws, r, "COMENTARIOS DE ESTUDIANTES")
        r += 1
        for course_label, comments in comments_by_course.items():
            r += 1
            style_section_header(ws, r, course_label)
            r += 1
            for comment in comments:
                cell = ws.cell(
                    row=r,
                    column=1,
                    value=(
                        comment.get("original_text", "")
                        if isinstance(comment, dict)
                        else comment
                    ),
                )
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws.row_dimensions[r].height = 40
                r += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"reporte_{detail['institutional_code']}"
        f"_{detail['period_code']}_{detail['evaluation_id']}.xlsx"
    )

    return buffer, filename


def teacher_streaming_response(buffer: io.BytesIO, filename: str) -> StreamingResponse:
    """Return a StreamingResponse for the teacher evaluation Excel report."""

    return StreamingResponse(
        buffer,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

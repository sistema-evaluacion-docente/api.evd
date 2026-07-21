"""
Shared styling constants and helpers for Excel exports.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

C_PRIMARY = "1A3A6B"
C_SECONDARY = "2E6DB4"
C_DIM_HDR = "D6E4F7"
C_ROW_ALT = "F4F8FF"
C_GREEN = "1B5E20"
C_ORANGE = "E65100"
C_RED = "B71C1C"
C_WHITE = "FFFFFF"

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def score_color(score):
    """Return a hex color code based on the score value."""

    if score is None:
        return "888888"
    if score >= 4.5:
        return C_GREEN
    if score >= 4.0:
        return "2E7D32"
    if score >= 3.5:
        return C_ORANGE
    return C_RED


def fill(hex_color):
    """Return a PatternFill object with the specified hex color."""

    return PatternFill("solid", fgColor=hex_color)


def bold_font(size=11, color="000000", white=False):
    """Return a Font object with bold styling and specified size and color."""

    return Font(bold=True, size=size, color=C_WHITE if white else color)


def thin_border():
    """Return a Border object with thin borders on all sides."""

    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    """Return an Alignment object for centered text with vertical centering and text wrapping."""

    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    """Return an Alignment object for left-aligned text with vertical centering and text wrapping."""

    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_title_row(ws, row, text, ncols=4):
    """Style a title row in the Excel worksheet."""

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=14, color=C_WHITE)
    cell.fill = fill(C_PRIMARY)
    cell.alignment = center()


def style_section_header(ws, row, text, ncols=4):
    """Style a section header row in the Excel worksheet."""

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=11, color=C_WHITE)
    cell.fill = fill(C_SECONDARY)
    cell.alignment = left()


def style_table_header(ws, row, cols):
    """Style a table header row in the Excel worksheet."""

    for col_idx, text in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = bold_font(white=True)
        cell.fill = fill(C_SECONDARY)
        cell.alignment = center()
        cell.border = thin_border()


def style_data_row(ws, row, values, alternate=False):
    """Style a data row in the Excel worksheet."""

    bg = C_ROW_ALT if alternate else C_WHITE
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.fill = fill(bg)
        cell.alignment = left()
        cell.border = thin_border()

"""
Routes for evaluation operations.
"""

import io
import os
import uuid

import openpyxl
from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from api.config import config
from api.controllers.evaluations import (
    EvaluationsController,
    get_evaluations_controller,
)
from api.database import get_db
from api.middlewares.auth import get_current_user, require_roles
from api.models.academic_period import AcademicPeriodModel
from api.models.department import DepartmentModel
from api.repositories.evaluations import (
    EvaluationsRepository,
    get_evaluations_repository,
)
from api.models.evaluation import EvaluationModel as EvaluationORM
from api.repositories.stats import StatsRepository, get_stats_repository
from api.repositories.users import UsersRepository, get_users_repository
from api.schemas.evaluation import (
    EvaluationDetailResponse,
    EvaluationListResponse,
    EvaluationStatusUpdate,
)
from api.schemas.evaluation_summary import (
    DimensionAveragesOut,
    EvaluationSummaryResponse,
    QuestionCatalogResponse,
    TeacherCommentsResponse,
    TeacherEvaluationDetailResponse,
    TeacherPeriodEvaluationsResponse,
)
from api.schemas.pagination import Pagination
from api.utils.dimensions import QUESTIONS
from api.schemas.response import ResponseSchema
from api.schemas.user import RoleName
from api.utils.evaluation_processor import process_evaluation
from api.utils.pdf_parser import parse_pdf
from api.utils.file_validation import validate_file_size

router = APIRouter(prefix="/evaluations", tags=["Evaluations"])


@router.post(
    "/upload",
    status_code=202,
    response_model=EvaluationDetailResponse,
    responses={
        400: {"model": ResponseSchema},
        409: {"model": ResponseSchema},
        422: {"model": ResponseSchema},
    },
)
async def upload_evaluation(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    current_user: dict = Depends(
        require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])
    ),
    db: Session = Depends(get_db),
    evaluations_repo: EvaluationsRepository = Depends(get_evaluations_repository),
    users_repo: UsersRepository = Depends(get_users_repository),
):
    """Upload a teacher evaluation PDF for a department.

    Parses the document immediately to validate period and department, then
    processes scores and comments in the background. Returns 202 with the
    evaluation record while processing continues.
    """

    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="File must be a PDF")

    pdf_bytes = await file.read()

    validate_file_size(pdf_bytes)

    if not pdf_bytes:
        raise HTTPException(
            status_code=400,
            detail="El archivo está vacío",
        )

    try:
        parsed = parse_pdf(pdf_bytes)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not parse PDF: {exc}")

    if not parsed.get("period_code"):
        raise HTTPException(
            status_code=422, detail="No se pudo extraer el periodo académico del PDF"
        )
    if not parsed.get("department_code"):
        raise HTTPException(
            status_code=422, detail="No se pudo extraer el departamento del PDF"
        )

    if not parsed.get("teachers"):
        raise HTTPException(
            status_code=422,
            detail="No se encontraron datos del docente en el PDF. Asegúrese de que se trate de un documento de evaluación docente de la UFPS.",
        )

    period = (
        db.query(AcademicPeriodModel)
        .filter(AcademicPeriodModel.code == parsed["period_code"])
        .first()
    )
    if not period:
        raise HTTPException(
            status_code=422,
            detail=f"Periodo académico '{parsed['period_code']}' no está registrado en el sistema",
        )

    department = (
        db.query(DepartmentModel)
        .filter(DepartmentModel.code == parsed["department_code"])
        .first()
    )
    if not department:
        raise HTTPException(
            status_code=422,
            detail=f"Departamento '{parsed['department_code']}' no está registrado en el sistema",
        )

    existing = await evaluations_repo.get_by_period_and_department(
        period.id, department.id
    )

    if existing and existing["active"] and existing["status"] == "COMPLETED":
        raise HTTPException(
            status_code=409,
            detail=(
                f"Una evaluación para el periodo '{parsed['period_code']}' "
                f"y este departamento ya existe"
            ),
        )

    if existing and existing["status"] in ("PROCESSING", "FAILED"):
        await evaluations_repo.delete(existing["id"])

    eval_dir = os.path.join(
        config.UPLOAD_DIR,
        "evaluations",
        parsed["period_code"],
        parsed["department_code"],
    )
    os.makedirs(eval_dir, exist_ok=True)

    ext = os.path.splitext(file.filename or "evaluation.pdf")[1] or ".pdf"
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(eval_dir, filename)

    with open(filepath, "wb") as f:
        f.write(pdf_bytes)

    user_record = await users_repo.get_by_uid(current_user["uid"])
    resolved_user_id = user_record["id"] if user_record else None

    evaluation = await evaluations_repo.create(
        user_id=resolved_user_id,
        academic_period_id=period.id,
        department_id=department.id,
        pdf_url=filepath,
        status="PROCESSING",
    )

    background_tasks.add_task(process_evaluation, evaluation["id"], parsed)

    return ResponseSchema(
        status=202,
        message="Carga de evaluación iniciada. Procesando en segundo plano.",
        data=evaluation,
        path="/evaluations/upload",
    )


@router.get(
    "/questions",
    response_model=QuestionCatalogResponse,
)
async def get_question_catalog(
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
):
    """Return the full catalog of evaluation questions with their code, text, and dimension."""

    return ResponseSchema(
        status=200,
        message="Question catalog retrieved successfully",
        data=QUESTIONS,
        path="/evaluations/questions",
    )


@router.get(
    "/",
    response_model=EvaluationListResponse,
    responses={403: {"description": "Forbidden"}},
)
async def get_all_evaluations(
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    search: str | None = Query(
        None,
        description="Search term for period name, teacher name, or course",
    ),
    period_id: int | None = Query(None, description="Filter by academic period ID"),
    department_id: int | None = Query(None, description="Filter by department ID"),
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Endpoint to list all evaluations with optional filters."""

    evaluations, total = await controller.get_all(
        page=page,
        limit=limit,
        search=search,
        period_id=period_id,
        department_id=department_id,
    )

    pages = (total + limit - 1) // limit if total else 0

    return ResponseSchema(
        status=200,
        message="Evaluations found",
        data=evaluations,
        pagination={"total": total, "page": page, "limit": limit, "pages": pages},
        path="/evaluations",
    )


@router.get(
    "/by-period/{period_id}",
    response_model=EvaluationDetailResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_evaluation_by_period(
    period_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Endpoint to get an evaluation by academic period ID."""

    evaluation = await controller.get_by_period(period_id)

    if not evaluation:
        return ResponseSchema(
            status=404,
            message="Evaluation not found for the given period",
            path=f"/evaluations/by-period/{period_id}",
        )

    return ResponseSchema(
        status=200,
        message="Evaluation found",
        data=evaluation,
        path=f"/evaluations/by-period/{period_id}",
    )


@router.get(
    "/{evaluation_id}",
    response_model=EvaluationDetailResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_evaluation_by_id(
    evaluation_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Endpoint to get an evaluation by ID."""

    evaluation = await controller.get_by_id(evaluation_id)

    if not evaluation:
        return ResponseSchema(
            status=404,
            message="Evaluation not found",
            path=f"/evaluations/{evaluation_id}",
        )

    return ResponseSchema(
        status=200,
        message="Evaluation found",
        data=evaluation,
        path=f"/evaluations/{evaluation_id}",
    )


@router.get(
    "/period/{period_id}/teachers",
    response_model=TeacherPeriodEvaluationsResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_teachers_by_period(
    period_id: int,
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(10, ge=1, le=100, description="Items per page"),
    search: str | None = Query(None, description="Search by teacher name or email"),
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Return all teachers with their average evaluation scores for a given academic period."""

    result = await controller.get_teachers_by_period(
        period_id, page=page, limit=limit, search=search
    )

    if not result:
        return ResponseSchema(
            status=404,
            message="Academic period not found or no evaluations exist for this period",
            path=f"/evaluations/period/{period_id}/teachers",
        )

    return ResponseSchema(
        status=200,
        message="Teacher evaluations retrieved successfully",
        data=result["teachers"],
        pagination=Pagination(
            total=result["teacher_count"],
            page=page,
            limit=limit,
            pages=result["pages"],
        ),
        path=f"/evaluations/period/{period_id}/teachers",
    )


@router.get(
    "/{evaluation_id}/summary",
    response_model=EvaluationSummaryResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_evaluation_summary(
    evaluation_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Return aggregated department statistics for an evaluation: department average,
    teacher ranking, and best/worst teacher identification."""

    summary = await controller.get_summary(evaluation_id)

    if not summary:
        return ResponseSchema(
            status=404,
            message="Evaluation not found",
            path=f"/evaluations/{evaluation_id}/summary",
        )

    return ResponseSchema(
        status=200,
        message="Summary generated successfully",
        data=summary,
        path=f"/evaluations/{evaluation_id}/summary",
    )


@router.get(
    "/{evaluation_id}/dimension-averages",
    response_model=DimensionAveragesOut,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_evaluation_dimension_averages(
    evaluation_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Return dimension-level averages aggregated across all groups for an evaluation."""

    dimensions = await controller.get_dimension_averages(evaluation_id)

    if dimensions is None:
        return ResponseSchema(
            status=404,
            message="Evaluation not found",
            path=f"/evaluations/{evaluation_id}/dimension-averages",
        )

    return ResponseSchema(
        status=200,
        message="Dimension averages retrieved successfully",
        data=dimensions,
        path=f"/evaluations/{evaluation_id}/dimension-averages",
    )


@router.get(
    "/{evaluation_id}/teachers/{teacher_id}/comments",
    response_model=TeacherCommentsResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_teacher_comments(
    evaluation_id: int,
    teacher_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Return comments grouped by course for a teacher within an evaluation."""

    result = await controller.get_teacher_comments(evaluation_id, teacher_id)

    if not result:
        return ResponseSchema(
            status=404,
            message="Evaluation or teacher not found",
            path=f"/evaluations/{evaluation_id}/teachers/{teacher_id}/comments",
        )

    return ResponseSchema(
        status=200,
        message="Comments retrieved successfully",
        data=result,
        path=f"/evaluations/{evaluation_id}/teachers/{teacher_id}/comments",
    )


@router.get(
    "/{evaluation_id}/teachers/{teacher_id}/export",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            }
        },
        403: {"description": "Forbidden"},
        404: {"model": ResponseSchema},
    },
)
async def export_teacher_evaluation(
    evaluation_id: int,
    teacher_id: int,
    include_comments: bool = Query(False, description="Include student comments"),
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
    db: Session = Depends(get_db),
    stats_repo: StatsRepository = Depends(get_stats_repository),
):
    """Download an Excel report for a teacher's evaluation detail."""

    detail = await controller.get_teacher_detail(evaluation_id, teacher_id)
    if not detail:
        raise HTTPException(status_code=404, detail="Evaluation or teacher not found")

    # Fetch teacher-vs-department comparison data
    comparison = None
    eval_record = (
        db.query(EvaluationORM).filter(EvaluationORM.id == evaluation_id).first()
    )
    if eval_record:
        comparison = await stats_repo.get_teacher_vs_department(
            teacher_id, eval_record.academic_period_id
        )

    dept_by_dim: dict[str, float | None] = {}
    dept_by_code: dict[str, float | None] = {}
    dept_name: str | None = None
    dept_overall: float | None = None

    if comparison:
        dept_name = comparison.get("department_name")
        dept_overall = comparison.get("department_overall_average")
        for dim in comparison["dimensions"]:
            dept_by_dim[dim["dimension"]] = dim["department_average"]
            for q in dim["questions"]:
                dept_by_code[q["code"]] = q["department_average"]

    comments_by_course: dict[str, list[str]] = {}
    if include_comments:
        comments_data = await controller.get_teacher_comments(evaluation_id, teacher_id)
        if comments_data:
            for c in comments_data["courses"]:
                key = f"{c['course_code']} - {c['course_name'] or ''}"
                comments_by_course[key] = c["comments"]

    # ── Paleta de colores ──────────────────────────────────────────────
    C_PRIMARY = "1A3A6B"  # azul oscuro — títulos principales
    C_SECONDARY = "2E6DB4"  # azul medio — cabeceras de sección
    C_DIM_HDR = "D6E4F7"  # azul muy claro — cabecera de tabla dimensión
    C_ROW_ALT = "F4F8FF"  # gris azulado — filas alternas
    C_GREEN = "1B5E20"  # verde oscuro — score alto
    C_ORANGE = "E65100"  # naranja — score medio
    C_RED = "B71C1C"  # rojo — score bajo
    C_WHITE = "FFFFFF"

    def score_color(score):
        if score is None:
            return "888888"
        if score >= 4.5:
            return C_GREEN
        if score >= 4.0:
            return "2E7D32"
        if score >= 3.5:
            return C_ORANGE
        return C_RED

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(size=11, color="000000", white=False):
        return Font(bold=True, size=size, color=C_WHITE if white else color)

    def thin_border():
        s = Side(style="thin", color="BBBBBB")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_title_row(ws, row, text, ncols=4):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=14, color=C_WHITE)
        cell.fill = fill(C_PRIMARY)
        cell.alignment = center()

    def style_section_header(ws, row, text, ncols=4):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, size=11, color=C_WHITE)
        cell.fill = fill(C_SECONDARY)
        cell.alignment = left()

    def style_table_header(ws, row, cols):
        for col_idx, text in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=col_idx, value=text)
            cell.font = bold_font(white=True)
            cell.fill = fill(C_SECONDARY)
            cell.alignment = center()
            cell.border = thin_border()

    def style_data_row(ws, row, values, alternate=False):
        bg = C_ROW_ALT if alternate else C_WHITE
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill(bg)
            cell.alignment = left()
            cell.border = thin_border()

    # ── Libro ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Docente"

    # Anchos de columna
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    r = 1  # puntero de fila

    # ── Título principal ───────────────────────────────────────────────
    style_title_row(ws, r, "REPORTE DE EVALUACIÓN DOCENTE — UFPS")
    ws.row_dimensions[r].height = 28
    r += 1
    r += 1  # fila en blanco

    # ── Info del docente ───────────────────────────────────────────────
    style_section_header(ws, r, "INFORMACIÓN DEL DOCENTE")
    r += 1

    info_rows = [
        ("Nombre", detail["name"] or "—"),
        ("Código", detail["institutional_code"]),
        ("Tipo contrato", detail["contract_type"] or "—"),
        ("Período", detail["period_code"] or "—"),
    ]
    for label, value in info_rows:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = bold_font()
        lc.fill = fill(C_DIM_HDR)
        lc.border = thin_border()
        lc.alignment = left()
        vc = ws.cell(row=r, column=2, value=value)
        vc.border = thin_border()
        vc.alignment = left()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    avg = detail["overall_average"]
    lc = ws.cell(row=r, column=1, value="Promedio docente")
    lc.font = bold_font()
    lc.fill = fill(C_DIM_HDR)
    lc.border = thin_border()
    lc.alignment = left()
    vc = ws.cell(row=r, column=2, value=f"{avg} / 5.0" if avg else "—")
    vc.font = Font(bold=True, color=score_color(avg))
    vc.border = thin_border()
    vc.alignment = left()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    lc = ws.cell(row=r, column=1, value="Promedio departamento")
    lc.font = bold_font()
    lc.fill = fill(C_DIM_HDR)
    lc.border = thin_border()
    lc.alignment = left()
    dept_avg_label = f"{dept_overall} / 5.0" if dept_overall else "N/D"
    if dept_name:
        dept_avg_label = (
            f"{dept_overall} / 5.0  ({dept_name})"
            if dept_overall
            else f"N/D  ({dept_name})"
        )
    vc = ws.cell(row=r, column=2, value=dept_avg_label)
    vc.font = Font(bold=True, color=score_color(dept_overall))
    vc.border = thin_border()
    vc.alignment = left()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1
    r += 1

    # ── Resumen por dimensión ──────────────────────────────────────────
    style_section_header(ws, r, "RESUMEN POR DIMENSIÓN")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Docente", "Departamento", "Diferencia"])
    r += 1
    for i, dim in enumerate(detail["dimensions"]):
        dim_avg = dim["average"]
        dept_dim_avg = dept_by_dim.get(dim["dimension"])
        diff = (
            round(dim_avg - dept_dim_avg, 2)
            if dim_avg is not None and dept_dim_avg is not None
            else None
        )
        diff_str = (
            f"+{diff}"
            if diff is not None and diff > 0
            else str(diff) if diff is not None else "N/D"
        )
        style_data_row(
            ws,
            r,
            [
                dim["dimension"],
                f"{dim_avg} / 5.0" if dim_avg is not None else "—",
                f"{dept_dim_avg} / 5.0" if dept_dim_avg is not None else "N/D",
                diff_str,
            ],
            alternate=i % 2 == 1,
        )
        ws.cell(row=r, column=2).font = Font(bold=True, color=score_color(dim_avg))
        ws.cell(row=r, column=3).font = Font(bold=True, color=score_color(dept_dim_avg))
        ws.cell(row=r, column=4).font = Font(
            bold=True, color=C_GREEN if (diff or 0) >= 0 else C_RED
        )
        r += 1
    r += 1

    # ── Fortalezas y oportunidades ─────────────────────────────────────
    sorted_dims = sorted(
        [d for d in detail["dimensions"] if d["average"] is not None],
        key=lambda d: d["average"],
        reverse=True,
    )
    style_section_header(ws, r, "FORTALEZAS — Top 2 dimensiones")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Promedio", "", ""])
    r += 1
    for dim in sorted_dims[:2]:
        style_data_row(ws, r, [dim["dimension"], f"{dim['average']} / 5.0", "", ""])
        ws.cell(row=r, column=2).font = Font(bold=True, color=C_GREEN)
        r += 1
    r += 1

    style_section_header(ws, r, "OPORTUNIDADES DE MEJORA — Top 2 dimensiones")
    r += 1
    style_table_header(ws, r, ["Dimensión", "Promedio", "", ""])
    r += 1
    for dim in sorted_dims[-2:]:
        style_data_row(ws, r, [dim["dimension"], f"{dim['average']} / 5.0", "", ""])
        ws.cell(row=r, column=2).font = Font(bold=True, color=C_RED)
        r += 1
    r += 1

    # ── Matriz de 22 ítems ─────────────────────────────────────────────
    style_title_row(ws, r, "MATRIZ DE EVALUACIÓN — 22 ÍTEMS")
    r += 1

    for dim in detail["dimensions"]:
        r += 1
        style_section_header(
            ws, r, f"{dim['dimension'].upper()}   |   Promedio: {dim['average']} / 5.0"
        )
        r += 1
        style_table_header(ws, r, ["Código", "Descripción", "Docente", "Departamento"])
        r += 1
        for i, q in enumerate(dim.get("questions", [])):
            dept_q_avg = dept_by_code.get(q["code"])
            style_data_row(
                ws,
                r,
                [
                    q["code"],
                    q["text"],
                    f"{q['score']} / 5.0" if q["score"] is not None else "—",
                    f"{dept_q_avg} / 5.0" if dept_q_avg is not None else "N/D",
                ],
                alternate=i % 2 == 1,
            )
            ws.cell(row=r, column=3).font = Font(
                bold=True, color=score_color(q["score"])
            )
            ws.cell(row=r, column=3).alignment = center()
            ws.cell(row=r, column=4).font = Font(
                bold=True, color=score_color(dept_q_avg)
            )
            ws.cell(row=r, column=4).alignment = center()
            r += 1

    # ── Comentarios (opcional) ─────────────────────────────────────────
    if include_comments and comments_by_course:
        r += 1
        style_title_row(ws, r, "COMENTARIOS DE ESTUDIANTES")
        r += 1
        for course_label, comments in comments_by_course.items():
            r += 1
            style_section_header(ws, r, course_label)
            r += 1
            for comment in comments:
                cell = ws.cell(row=r, column=1, value=comment)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
                ws.row_dimensions[r].height = 40
                r += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_{detail['institutional_code']}_{detail['period_code']}_{evaluation_id}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get(
    "/{evaluation_id}/teachers/{teacher_id}",
    response_model=TeacherEvaluationDetailResponse,
    responses={403: {"description": "Forbidden"}, 404: {"model": ResponseSchema}},
)
async def get_teacher_evaluation_detail(
    evaluation_id: int,
    teacher_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Return per-course and per-dimension scores for a teacher within an evaluation."""

    detail = await controller.get_teacher_detail(evaluation_id, teacher_id)

    if not detail:
        return ResponseSchema(
            status=404,
            message="Evaluation or teacher not found",
            path=f"/evaluations/{evaluation_id}/teachers/{teacher_id}",
        )

    return ResponseSchema(
        status=200,
        message="Teacher evaluation detail retrieved successfully",
        data=detail,
        path=f"/evaluations/{evaluation_id}/teachers/{teacher_id}",
    )


@router.get(
    "/{evaluation_id}/export",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            }
        },
        403: {"description": "Forbidden"},
        404: {"model": ResponseSchema},
    },
)
async def export_evaluation(
    evaluation_id: int,
    _=Depends(require_roles([RoleName.ADMIN, RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Download an Excel file with the department evaluation summary."""

    summary = await controller.get_summary(evaluation_id)

    if not summary:
        raise HTTPException(status_code=404, detail="Evaluation not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Evaluación Docente"

    ws.append(
        ["Ranking", "Nombre", "Código", "Tipo Contrato", "Grupos Evaluados", "Promedio"]
    )

    for item in summary["ranking"]:
        ws.append(
            [
                item["rank"],
                item["name"] or "",
                item["institutional_code"],
                item["contract_type"] or "",
                item["group_count"],
                item["overall_average"],
            ]
        )

    ws.append([])
    ws.append(["", "Promedio Departamento", "", "", "", summary["department_average"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"evaluacion_{summary['period_code']}_{evaluation_id}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.patch(
    "/{evaluation_id}/status",
    response_model=EvaluationDetailResponse,
    responses={
        400: {"model": ResponseSchema},
        403: {"description": "Forbidden"},
        404: {"model": ResponseSchema},
    },
)
async def update_evaluation_status(
    evaluation_id: int,
    payload: EvaluationStatusUpdate,
    current_user=Depends(get_current_user),
    _=Depends(require_roles([RoleName.DIRECTOR_DE_DEPARTAMENTO])),
    controller: EvaluationsController = Depends(get_evaluations_controller),
):
    """Activate or deactivate an evaluation."""

    evaluation = await controller.update_status(
        evaluation_id, payload.active, current_user
    )

    if not evaluation:
        return ResponseSchema(
            status=404,
            message="Evaluation not found",
            path=f"/evaluations/{evaluation_id}/status",
        )

    action = "activated" if payload.active else "deactivated"

    return ResponseSchema(
        status=200,
        message=f"Evaluation {action} successfully",
        data=evaluation,
        path=f"/evaluations/{evaluation_id}/status",
    )
